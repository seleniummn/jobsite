import openpyxl
book =openpyxl.load_workbook("C:\PythonExcel\PythonDemo.xlsx")
# set control of the sheet to the active worksheet in the excel spreadsheet
sheet =book.active
# create an empty dictionary value to insert the data into
Dict = {}
# set the sheet row and cell we are obtaining the data from
cell =sheet.cell(row=1, column=2)
# take the value from cells in the row and column details set above and print
#print(cell.value)
# output from the above = 'firstname' as we have selected the title row

# write data INTO the excel sheet at the specified row and column cell
sheet.cell(row=3, column=2).value = "TesterMN"
#Save the newly added data
book.save('C:\PythonExcel\PythonDemo.xlsx')

# print the input we have just added above to check it
print(sheet.cell(row=3, column=2).value)

# print number of rows of data in sheet
print(sheet.max_row)

# print number of columns of data in sheet
print(sheet.max_column)

# print the data value at a defined cell
print(sheet['A5'].value)

# #obtain and print full contents of the sheet
# for i in range(1,sheet.max_row+1):
#     for j in range(1,sheet.max_column+1):
#         print(sheet.cell(row=i, column =j).value)

# obtain row values from row 1 to the maximum available row (using 'sheet.max_row+1')
#if there were 15 rows, using 'range(1,15)' would only obtain 14 rows, hence '(2 = starting row (first data), sheet.max_row+1 = end row)'
for i in range(2,sheet.max_row+1):  # to get rows
    # we then run through each row, obtaining the value for column 1 ('NAME' - 'TestCase1','TestCase2', etc)
    # here, 'i' in 'row = i' is the value set above in 'for i in range'
    print(sheet.cell(row =i,column=1).value)
    # ADD CONDITION to test only one area of the spreadsheet data by checking for existence of 'TestCase2' value
    # row 'a' starts data at row '2' (first row below headers) of sheet
    if sheet.cell(row =i,column=1).value == "TestCase5":


    #obtain row values from column 2 to the maximum available column (using 'sheet.max_column+1')
        for j in range(2,sheet.max_column+1):# to get columns
            #add further data into the dictionary
            #Dict["lastname"]="Wolfeschlegelsteinhausenbergerdorff"
            #obtain data from the existing excel sheet, then place it into the dictionary
            #row 1 contains all header titles such as 'first name', hence 'row = 1'
            Dict[sheet.cell(row=1, column=j).value]= sheet.cell(row=i, column=j).value

#print the contents of the dictionary
print(Dict)










