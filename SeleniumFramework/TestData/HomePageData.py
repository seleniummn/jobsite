

class HomePageData:


    #data dictionary to pass to the 'test_homepage' test section
    test_HomePage_data = [{"firstname": "Matt", "lastname": "Smith", "gender": "Male"}, {"firstname": "Jane", "lastname": "Smith", "gender": "Female"}]


import openpyxl


class HomePageData:
    test_HomePage_data = [{"firstname": "Matt", "lastname": "Smith", "gender": "Male"},
                          {"firstname": "Jane", "lastname": "Smith", "gender": "Female"}]

    #below is for using excel data import in the place of the above data line
    #static method declared, so 'self' is not required prior to 'test_case_name'
    @staticmethod
    def getTestData(test_case_name):
        Dict = {}
        book = openpyxl.load_workbook("C:\\PythonExcel\\PythonDemo.xlsx")
        sheet = book.active
        for i in range(1, sheet.max_row + 1):  # to get rows
            if sheet.cell(row=i, column=1).value == test_case_name:

                for j in range(2, sheet.max_column + 1):  # to get columns
                    # Dict["lastname"]="shetty
                    Dict[sheet.cell(row=1, column=j).value] = sheet.cell(row=i, column=j).value
        return[Dict]